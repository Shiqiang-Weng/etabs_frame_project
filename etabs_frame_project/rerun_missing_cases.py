#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rerun missing analysis cases after the full 30k pipeline completes."""

from __future__ import annotations

import csv
import io
import json
import re
import shutil
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional

_CASE_PATTERN = re.compile(r"case_(\d+)", re.IGNORECASE)


@dataclass(frozen=True)
class RerunResult:
    case_id: int
    status: str
    message: Optional[str] = None


def _find_missing_report(project_root: Path, output_root: Path) -> Optional[Path]:
    prefix = "[缺失重跑]"
    base_name = "analysis_data_missing_report"
    extension_priority = [".csv", ".md", ".txt", ".xlsx", ".xls", ""]
    candidates = [
        f"{base_name}.csv",
        f"{base_name}.md",
        f"{base_name}.txt",
        f"{base_name}.xlsx",
        f"{base_name}.xls",
    ]

    def _pick_from_glob(paths: List[Path]) -> Optional[Path]:
        if not paths:
            return None
        def _priority(path: Path) -> int:
            suffix = path.suffix.lower()
            return extension_priority.index(suffix) if suffix in extension_priority else len(extension_priority)
        return sorted(paths, key=lambda p: (_priority(p), p.name))[0]

    extra_backup_paths = [
        Path(r"D:\GNN_data\data\output\raw\etabs_script_output_frame"),
    ]
    search_roots = [output_root, project_root, *extra_backup_paths]
    for root in search_roots:
        for name in candidates:
            path = root / name
            if path.exists():
                print(f"{prefix} 缺失报告路径: {path} (suffix={path.suffix or '<none>'})")
                return path

        glob_matches = sorted(root.glob(f"{base_name}*"))
        if glob_matches:
            selected = _pick_from_glob(glob_matches)
            if selected:
                sample = [str(p) for p in glob_matches[:5]]
                print(f"{prefix} 缺失报告路径: {selected} (suffix={selected.suffix or '<none>'})")
                print(f"{prefix} glob 命中: {sample}")
                return selected

    attempted = ", ".join(candidates + [f"{base_name}*"])
    attempted_roots = ", ".join(str(root) for root in search_roots)
    print(
        f"{prefix} 未找到 analysis_data_missing_report*（已尝试：{attempted}；"
        f"search_root={attempted_roots}）"
    )
    return None


def _read_text_with_fallback(report_path: Path, encodings: List[str], final_errors: Optional[str] = None) -> str:
    for encoding in encodings[:-1]:
        try:
            return report_path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    last_encoding = encodings[-1] if encodings else "utf-8"
    try:
        return report_path.read_text(encoding=last_encoding)
    except UnicodeDecodeError:
        return report_path.read_text(encoding=last_encoding, errors=final_errors or "ignore")


def _normalize_field_name(name: str) -> str:
    return re.sub(r"[\s_]+", "", name.strip().lower())


def _parse_case_id_from_text(text: str) -> Optional[int]:
    if not text:
        return None
    match = _CASE_PATTERN.search(text)
    if match:
        return int(match.group(1))
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _extract_case_ids_from_text(text: str) -> List[int]:
    ids: set[int] = set()
    for match in _CASE_PATTERN.finditer(text):
        ids.add(int(match.group(1)))
    for match in re.finditer(r"\b\d+\b", text):
        try:
            ids.add(int(match.group(0)))
        except ValueError:
            continue
    return sorted(ids)


def _extract_case_ids_from_csv(report_path: Path) -> List[int]:
    text = _read_text_with_fallback(report_path, ["utf-8-sig", "gbk", "utf-8"], final_errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []
    normalized = {_normalize_field_name(name): name for name in fieldnames if name}
    case_field = None
    for candidate in ("case", "caseid", "case_id"):
        normalized_name = candidate.replace("_", "")
        if normalized_name in normalized:
            case_field = normalized[normalized_name]
            break

    ids: set[int] = set()
    if case_field:
        for row in reader:
            raw = (row.get(case_field) or "").strip()
            case_id = _parse_case_id_from_text(raw)
            if case_id is not None:
                ids.add(case_id)
    else:
        for row in csv.reader(io.StringIO(text)):
            row_text = " ".join(str(value) for value in row if value is not None)
            for match in _CASE_PATTERN.finditer(row_text):
                ids.add(int(match.group(1)))

    return sorted(ids)


def _extract_case_ids_from_markdown(report_path: Path) -> List[int]:
    text = _read_text_with_fallback(report_path, ["utf-8-sig", "utf-8", "gbk"], final_errors="ignore")
    return _extract_case_ids_from_text(text)


def _extract_case_ids_from_excel(report_path: Path) -> List[int]:
    try:
        import openpyxl
    except ImportError:
        print("[缺失重跑] 未安装 openpyxl，无法解析 Excel 报告，请导出为 CSV/Markdown。")
        return []

    workbook = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    sheet = workbook.active
    ids: set[int] = set()

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    case_col_idx = None
    if header_row:
        for idx, value in enumerate(header_row):
            if value is None:
                continue
            normalized = _normalize_field_name(str(value))
            if normalized in ("case", "caseid"):
                case_col_idx = idx
                break

    if case_col_idx is not None:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if case_col_idx >= len(row):
                continue
            cell_value = row[case_col_idx]
            case_id = _parse_case_id_from_text("" if cell_value is None else str(cell_value))
            if case_id is not None:
                ids.add(case_id)
    else:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                for match in _CASE_PATTERN.finditer(str(cell)):
                    ids.add(int(match.group(1)))

    return sorted(ids)


def _extract_case_ids(report_path: Path) -> List[int]:
    suffix = report_path.suffix.lower()
    if suffix == ".csv":
        return _extract_case_ids_from_csv(report_path)
    if suffix in (".md", ".txt", ""):
        return _extract_case_ids_from_markdown(report_path)
    if suffix in (".xlsx", ".xlsm"):
        return _extract_case_ids_from_excel(report_path)
    if suffix == ".xls":
        print("[缺失重跑] 检测到 .xls 报告，请导出为 CSV/Markdown 以便解析。")
        return []
    return _extract_case_ids_from_markdown(report_path)


def _remove_done_flag(done_flag_path: Path) -> None:
    if done_flag_path.exists():
        done_flag_path.unlink()


def _remove_analysis_data(case_output_dir: Path) -> None:
    analysis_dir = case_output_dir / "data_extraction" / "analysis_data"
    if analysis_dir.exists():
        shutil.rmtree(analysis_dir, ignore_errors=True)


def _write_rerun_log(output_root: Path, results: Iterable[RerunResult]) -> Path:
    log_path = output_root / "missing_cases_rerun_log.json"
    payload = {
        "timestamp": time.time(),
        "results": [
            {"case_id": r.case_id, "status": r.status, "message": r.message} for r in results
        ],
    }
    log_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return log_path


def rerun_missing_cases(
    *,
    design_cases: List,
    run_case_fn: Callable,
    options,
    project_root: Path,
    output_root: Path,
    get_done_flag_path: Callable[[int], Path],
    get_case_output_dir: Callable[[int], Path],
) -> None:
    """
    Rerun cases listed in analysis_data_missing_report.csv after the full run completes.

    This function is intentionally defensive: missing files or individual case failures
    do not abort the entire pipeline.
    """
    prefix = "[缺失重跑]"
    report_path = _find_missing_report(project_root, output_root)
    if report_path is None:
        print(f"{prefix} 未找到 analysis_data_missing_report*，跳过缺失重跑。")
        return

    try:
        print(f"{prefix} 缺失报告: {report_path} (suffix={report_path.suffix or '<none>'})")
        missing_ids = _extract_case_ids(report_path)
    except Exception as exc:  # noqa: BLE001
        print(f"{prefix} 读取缺失报告失败: {exc}，跳过缺失重跑。")
        return

    if not missing_ids:
        print(f"{prefix} 缺失报告为空或无有效 case_id，跳过缺失重跑。")
        return

    print(f"{prefix} 缺失报告路径: {report_path}")
    print(f"{prefix} 需要重跑的案例数: {len(missing_ids)}")

    case_lookup: Dict[int, object] = {case.case_id: case for case in design_cases}
    results: List[RerunResult] = []
    total = len(missing_ids)
    success_count = 0
    failed_count = 0
    skipped_count = 0

    for idx, case_id in enumerate(missing_ids, start=1):
        design_case = case_lookup.get(case_id)
        if design_case is None:
            results.append(RerunResult(case_id=case_id, status="skipped", message="case_id not found"))
            skipped_count += 1
            continue

        try:
            done_flag = get_done_flag_path(case_id)
            _remove_done_flag(done_flag)
            _remove_analysis_data(get_case_output_dir(case_id))

            run_case_fn(design_case, idx, total, options)
            results.append(RerunResult(case_id=case_id, status="success"))
            success_count += 1
        except Exception as exc:  # noqa: BLE001
            results.append(RerunResult(case_id=case_id, status="failed", message=str(exc)))
            failed_count += 1

    log_path = _write_rerun_log(Path(output_root), results)
    failed_ids = [r.case_id for r in results if r.status == "failed"]
    print(
        f"{prefix} 完成: total={total}, success={success_count}, "
        f"failed={failed_count}, skipped={skipped_count}"
    )
    if failed_ids:
        print(f"{prefix} 失败 case_id: {failed_ids}")
    print(f"{prefix} 日志已写入: {log_path}")
